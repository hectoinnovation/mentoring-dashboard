from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import config


def _fill(hex_color: str) -> PatternFill:
    if hex_color == "00000000":
        return PatternFill(fill_type=None)
    return PatternFill("solid", fgColor=hex_color[2:] if hex_color.startswith("FF") else hex_color)


def _font(bold=False, size=9, color="000000", name=None) -> Font:
    return Font(
        name=name or config.FONT_NAME,
        bold=bold,
        size=size,
        color=color if len(color) == 6 else color[2:],
    )


def _align(h="center", v="center", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(style="thin", color="000000") -> Border:
    side = Side(border_style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _border_outline(ws, min_row, max_row, min_col, max_col, style="thin"):
    """셀 범위의 바깥쪽 테두리만 그린다."""
    thick = Side(border_style=style, color="000000")
    none_side = Side(border_style=None)

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            top = thick if cell.row == min_row else none_side
            bottom = thick if cell.row == max_row else none_side
            left = thick if cell.column == min_col else none_side
            right = thick if cell.column == max_col else none_side
            cell.border = Border(top=top, bottom=bottom, left=left, right=right)


def apply_cell(ws, row, col, value, bg_color, fg_color, bold, size, h_align="center", wrap=True, merge_end_col=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(bg_color)
    cell.font = _font(bold=bold, size=size, color=fg_color)
    cell.alignment = _align(h=h_align, v="center", wrap=wrap)
    cell.border = _border("thin")

    if merge_end_col and merge_end_col > col:
        end_letter = get_column_letter(merge_end_col)
        start_letter = get_column_letter(col)
        ws.merge_cells(f"{start_letter}{row}:{end_letter}{row}")

    return cell


def apply_label_col(ws, row, text, size=10, bold=False):
    cell = ws.cell(row=row, column=config.COL_LABEL, value=text)
    cell.font = _font(bold=bold, size=size)
    cell.alignment = _align(h="left", v="center")


def apply_connector_row(ws, row, min_col, max_col):
    """본부/실/팀 사이 연결선 역할 행 - 하단 중앙에 수직선 표현."""
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _fill("00000000")
