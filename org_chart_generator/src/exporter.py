import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from typing import List
import config
from src.org_tree import Department, Division, Team, Member


def _fill(hex6: str) -> PatternFill:
    if not hex6:
        return PatternFill(fill_type=None)
    argb = hex6 if len(hex6) == 8 else "FF" + hex6
    return PatternFill("solid", fgColor=argb)


def _font(bold=False, size=9, color="000000") -> Font:
    return Font(name=config.FONT_NAME, bold=bold, size=size, color=color)


def _align(h="center", v="center", wrap=True) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _side(style="thin") -> Side:
    return Side(border_style=style, color="000000")


def _border_all(style="thin") -> Border:
    s = _side(style)
    return Border(left=s, right=s, top=s, bottom=s)


def _merge(ws, r1, c1, r2, c2):
    if r1 == r2 and c1 == c2:
        return
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _style_range(ws, r1, c1, r2, c2, bg6=None, fg6="000000", bold=False, size=9,
                 h="center", wrap=True, border="thin"):
    fill = _fill(bg6) if bg6 else PatternFill(fill_type=None)
    fnt = _font(bold=bold, size=size, color=fg6)
    aln = _align(h=h, v="center", wrap=wrap)
    brd = _border_all(border)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            cell.font = fnt
            cell.alignment = aln
            cell.border = brd


def _box(ws, r1, c1, r2, c2, value, bg6=None, fg6="000000", bold=False, size=9,
         h="center", wrap=True, border="thin", merge=True):
    """범위를 스타일링하고, merge하고, r1/c1에 값을 씀."""
    _style_range(ws, r1, c1, r2, c2, bg6=bg6, fg6=fg6, bold=bold, size=size,
                 h=h, wrap=wrap, border=border)
    if merge:
        _merge(ws, r1, c1, r2, c2)
    ws.cell(row=r1, column=c1, value=value)


def _name_rank_row(ws, row, col_start, col_span, name, rank, bg6=None):
    """이름(좌) + 직급(우) 형태로 한 행 그리기 - 원본 스타일."""
    half = col_span // 2
    name_end = col_start + half - 1
    rank_start = col_start + half
    rank_end = col_start + col_span - 1

    _box(ws, row, col_start, row, name_end, name or "-",
         bg6=bg6, fg6="000000", size=config.FONT_MEMBER)
    _box(ws, row, rank_start, row, rank_end, rank or "-",
         bg6=bg6, fg6="000000", size=config.FONT_RANK)


def _draw_ceo_box(ws, ceo_name: str, ceo_rank: str, col_start: int, col_end: int):
    R_NAME = config.ROW_NUM_CEO_NAME
    half_span = (col_end - col_start + 1) // 2
    if half_span < 1:
        half_span = col_end - col_start + 1

    _box(ws, R_NAME - 1, col_start, R_NAME - 1, col_end, "CEO",
         bg6="222222", fg6="FFFFFF", bold=True, size=config.FONT_CEO)
    _name_rank_row(ws, R_NAME, col_start, col_end - col_start + 1,
                   ceo_name, ceo_rank, bg6=None)


def _draw_dept_box(ws, dept: Department, col_start: int, col_end: int):
    R1 = config.ROW_NUM_DEPT
    R2 = config.ROW_NUM_DEPT_LEADER
    span = col_end - col_start + 1

    _box(ws, R1, col_start, R1, col_end, dept.name,
         bg6="A5A5A5", fg6="FFFFFF", bold=True, size=config.FONT_DEPT)

    # 리더 행: 이름 | 직급
    _name_rank_row(ws, R2, col_start, span,
                   dept.leader_name, dept.leader_rank, bg6="A5A5A5")
    # 리더행 글자색 흰색으로 덮어씌우기
    half = span // 2
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=R2, column=c)
        cell.font = _font(bold=False, size=config.FONT_MEMBER, color="FFFFFF")


def _draw_div_box(ws, div: Division, col_start: int, col_end: int):
    R1 = config.ROW_NUM_DIV
    R2 = config.ROW_NUM_DIV_LEADER
    span = col_end - col_start + 1

    div_name = div.name if div.name else "-"
    _box(ws, R1, col_start, R1, col_end, div_name,
         bg6="FF6114", fg6="FFFFFF", bold=True, size=config.FONT_DIV)

    _name_rank_row(ws, R2, col_start, span,
                   div.leader_name, div.leader_rank, bg6="FF6114")
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=R2, column=c)
        cell.font = _font(bold=False, size=config.FONT_MEMBER, color="FFFFFF")


def _draw_team_box(ws, team: Team, col_start: int, col_end: int):
    _box(ws, config.ROW_NUM_TEAM, col_start, config.ROW_NUM_TEAM, col_end,
         team.name, bg6="FFAE81", fg6="000000", bold=True, size=config.FONT_TEAM)


def _draw_members(ws, team: Team, col_start: int):
    R_START = config.ROW_NUM_MEMBER_START
    cols_pp = config.COLS_PER_PERSON
    name_cols = cols_pp // 2
    col_end = col_start + cols_pp - 1

    for i, member in enumerate(team.members):
        row = R_START + i
        is_leader = bool(member.role)
        name_bg = "FFF2CC" if is_leader else None
        rank_bg = "FFF2CC" if is_leader else None

        n1, n2 = col_start, col_start + name_cols - 1
        r1, r2 = col_start + name_cols, col_end

        _box(ws, row, n1, row, n2, member.name,
             bg6=name_bg, fg6="000000", size=config.FONT_MEMBER)
        _box(ws, row, r1, row, r2, member.rank,
             bg6=rank_bg, fg6="000000", size=config.FONT_RANK)


def _set_label(ws, row, text, bold=False, size=10):
    cell = ws.cell(row=row, column=config.COL_LABEL, value=text)
    cell.font = _font(bold=bold, size=size)
    cell.alignment = _align(h="left", v="center", wrap=True)
    cell.border = _border_all("thin")


def generate(departments: List[Department], company_name: str, output_path: str,
             ceo_name: str = "대표이사", ceo_rank: str = ""):
    wb = Workbook()
    ws = wb.active
    ws.title = "조직도"
    ws.sheet_view.showGridLines = False

    # ── 총 컬럼 수 계산 ──
    total_content_cols = sum(d.col_span for d in departments)
    total_end_col = config.COL_CONTENT_START + total_content_cols - 1

    # ── 열 너비 ──
    ws.column_dimensions["A"].width = config.COL_LABEL_WIDTH
    ws.column_dimensions["B"].width = config.COL_SEP_WIDTH
    for i in range(total_content_cols):
        ws.column_dimensions[get_column_letter(config.COL_CONTENT_START + i)].width = config.COL_NAME_WIDTH

    # ── 행 높이 ──
    row_heights = {
        1: 16.5,
        config.ROW_NUM_TITLE: config.ROW_TITLE,
        3: 19.5, 4: 22.5, 5: 22.5,
        config.ROW_NUM_CEO_LABEL: 22.5,
        config.ROW_NUM_CEO_NAME: 24.75,
        8: 32.25, 9: 33.0,
        10: config.ROW_GAP_SM, 11: config.ROW_GAP_SM,
        config.ROW_NUM_DEPT: config.ROW_DEPT,
        config.ROW_NUM_DEPT_LEADER: config.ROW_DEPT,
        config.ROW_NUM_GAP: config.ROW_GAP_MD,
        15: config.ROW_GAP_LG,
        config.ROW_NUM_DIV: config.ROW_DIV,
        config.ROW_NUM_DIV_LEADER: config.ROW_DIV,
        config.ROW_NUM_GAP2: 21.75,
        19: 22.5,
        config.ROW_NUM_TEAM: config.ROW_TEAM,
    }
    for r, h in row_heights.items():
        ws.row_dimensions[r].height = h

    max_members = max(
        (len(t.members) for d in departments for div in d.divisions for t in div.teams),
        default=1,
    )
    for i in range(max_members):
        ws.row_dimensions[config.ROW_NUM_MEMBER_START + i].height = config.ROW_MEMBER

    # ── 레이블 열 (A열) ──
    _set_label(ws, config.ROW_NUM_CEO_LABEL, "대표이사")
    _set_label(ws, config.ROW_NUM_DEPT, "본부/부", bold=True)
    _set_label(ws, config.ROW_NUM_DEPT_LEADER, "본부장/\n부서장")
    _set_label(ws, config.ROW_NUM_DIV, "실")
    _set_label(ws, config.ROW_NUM_DIV_LEADER, "실장")
    _set_label(ws, config.ROW_NUM_TEAM, "팀", bold=True)
    _set_label(ws, config.ROW_NUM_MEMBER_START, "팀장")

    # ── 회사 제목 ──
    title_cell = ws.cell(row=config.ROW_NUM_TITLE,
                         column=config.COL_CONTENT_START, value=company_name)
    title_cell.font = Font(name=config.FONT_NAME, bold=True, size=config.FONT_TITLE)
    title_cell.alignment = _align(h="center")
    if total_end_col > config.COL_CONTENT_START:
        _merge(ws, config.ROW_NUM_TITLE, config.COL_CONTENT_START,
               config.ROW_NUM_TITLE, total_end_col)

    # ── CEO 박스 (전체 너비) ──
    _draw_ceo_box(ws, ceo_name, ceo_rank,
                  config.COL_CONTENT_START, total_end_col)

    # ── 본부 / 실 / 팀 / 구성원 ──
    cur_col = config.COL_CONTENT_START

    for dept in departments:
        dept_col_start = cur_col
        dept_col_end = cur_col + dept.col_span - 1

        _draw_dept_box(ws, dept, dept_col_start, dept_col_end)

        for div in dept.divisions:
            div_col_start = cur_col
            div_col_end = cur_col + div.col_span - 1

            _draw_div_box(ws, div, div_col_start, div_col_end)

            for team in div.teams:
                team_col_start = cur_col
                team_col_end = cur_col + team.col_span - 1

                _draw_team_box(ws, team, team_col_start, team_col_end)
                _draw_members(ws, team, team_col_start)

                cur_col += team.col_span

    # ── 인쇄 설정 ──
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    wb.save(output_path)
    print("저장 완료:", output_path)
